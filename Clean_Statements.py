from bs4 import BeautifulSoup
import requests
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from openpyxl import load_workbook
from datetime import datetime
import re
import os
import glob
import openpyxl




class Clean_Statements():
    
    def __init__(self, statement, statement_required, date, company):
        

        self.statement = statement
        self.typestatement = statement_required
        self.date = date
        self.company = company
        self.path = f"====WRITE THE PATH FOR THE UNCLEANED STATEMENTS===={self.company}/"+str(self.typestatement)+f"{self.company}.xlsx"
    
    
    def download(self, df):
        
        
        if os.path.isdir(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}') == True:
            
            if os.path.isfile(self.path) == True:
                try:
                    with pd.ExcelWriter(self.path ,engine='openpyxl', mode='a') as writer:
                        df.to_excel(writer, sheet_name=str(self.date))
                except FileNotFoundError:
                    print("No FIle")
                    
            else:
                try:
                    with pd.ExcelWriter(self.path ,engine='openpyxl', mode='w') as writer:
                        df.to_excel(writer, sheet_name=str(self.date))
                except FileNotFoundError:
                    print("No FIle")
                
        else:
        
            try:
                os.mkdir(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}')
                with pd.ExcelWriter(self.path ,engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name=str(self.date))
            except FileNotFoundError:
                print("Not worked")
    
    
    
    
    def convert_data(self, statement):
        '''It takes the statement and edits it'''
        
        statement = statement.replace('[$]\s(?=[0-9])', "", regex=True)
        statement = statement.replace('\\n', "", regex=True)
        statement = statement.replace('[:]', "", regex=True)

        statement = statement.replace(",", "", regex=True)
        statement = statement.replace(" ", "0", regex=True)

        statement = statement.replace("[(]", "-", regex=True)
        statement = statement.replace("[)]", "", regex=True)
        statement.reset_index(inplace = True)
        statement = statement.replace("\[(.*?)\]", "", regex=True)
        statement = statement.replace("\((.*?)\)", "", regex=True)
        statement = statement.replace('[:]', "", regex=True)
        statement['Account'] = statement['Account'].str.strip()
       
        statement.set_index('Account', inplace=True)
        statement = statement.replace('[$ ]', "", regex=True)
        
        for j in statement.columns:    

            number = []
            for i in statement[j]:
                if i == '\xa0':
                    number.append(np.nan)
                else:
                    number.append(pd.to_numeric(i))
            statement[j] = number
            
        statement.reset_index(inplace = True)
        for i in range(len(statement.index)):
            if "Basic" == statement.iloc[i]['Account']:

                if  statement.iloc[i][1]<10000:
                    statement.loc[i, 'Account'] = "Basic Shares"
                    pass
                else:
                    statement.loc[i, 'Account'] = "Basic Dollars_per_share"
                    pass

            elif "Diluted" == statement.iloc[i]['Account']:

                if  statement.iloc[i][1]<10000:
                    statement.loc[i, 'Account'] = "Diluted Shares"
                    pass
                else:
                    statement.loc[i, 'Account'] = "Diluted Dollars_per_share"
                    pass
        statement.set_index('Account', inplace=True)
        return statement
        
    def get_sheets(self):
        
        
        if os.path.isdir(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}') == True:
            
            if os.path.isfile(self.path) == True:
                
                files = glob.iglob(self.path)
                sheet = []
                for fileName in files:
                    wb = openpyxl.load_workbook(fileName)
                    sheet.append(wb.sheetnames)
                
                if self.date in sheet:
                    print("This sheet has already been cleaned")
                    
                else:
                    return self.date
                
                    
            else:
                return False
                
        else:
        
            os.mkdir(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}')
            return False

        
    def merging_data(self, statement):
        
        path = f"====PATH FOR THE CLEANED STATEMENTS====/{self.company}/Merged_"+str(self.typestatement)+f"{self.company}.xlsx"
        
        print(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}')
        if os.path.isdir(f'====PATH FOR THE CLEANED STATEMENTS===={self.company}') == True:
            
            if os.path.isfile(path) == True:
                
                print("Cleaned_Reports/"+str(self.company)+"/"+str(self.typestatement)+str(self.company)+".xlsx")
                new_statement = pd.read_excel("Cleaned_Reports/"+str(self.company)+"/"+str(self.typestatement)+str(self.company)+".xlsx", index_col=0, usecols=[1], sheet_name = str(self.date))
                new_statement.reset_index(inplace=True)
                merged_statement = pd.read_excel(r"Cleaned_Reports/"+str(self.company)+"/Merged_"+str(self.typestatement)+str(self.company)+".xlsx", index_col=0)
                merged_statement.reset_index(inplace=True)
                if new_statement.columns[0] not in merged_statement.columns:

                    joined_statement = pd.concat([merged_statement, new_statement],1, join='outer')
                    
                    joined_statement.set_index("Account", inplace=True)
                    joined_statement.columns = [datetime.strptime(i, '%m/%Y') for i in joined_statement.columns]
                    joined_statement.columns = [datetime.strftime(i, '%m/%Y') for i in joined_statement.columns]
                    

                    try:
                        with pd.ExcelWriter(path ,engine='openpyxl', mode='w') as writer:
                            joined_statement.to_excel(writer, sheet_name=str(self.company))
                    except FileNotFoundError:
                        print("No FIle")
                else:
                    print("Quarter is already merged")

            else:  

                new_statement = pd.read_excel("Cleaned_Reports/"+str(self.company)+"/"+str(self.typestatement)+str(self.company)+".xlsx", index_col=0, usecols=[0,1], sheet_name = str(self.date))
                try:
                    with pd.ExcelWriter(path ,engine='openpyxl', mode='w') as writer:
                        new_statement.to_excel(writer, sheet_name=str(self.company))
                except FileNotFoundError:
                    print("Not worked")
        
        
    def Cleaning(self):
        
        
        if self.typestatement == 'IS':
            
            sheet = self.get_sheets()
            
            '''Create an Excel Sheet Because there is no in the directory'''

            income_statement = self.statement
            print(income_statement.columns)

            income_statement = self.convert_data(income_statement)

            self.download(income_statement)
            
            '''Now we have to join the dataframes in order to make a time series prediction'''
            
            
            self.merging_data(income_statement)
                
                
            
        elif self.typestatement == 'BS':
            sheet = self.get_sheets()
            
            '''Create an Excel Sheet Because there is no in the directory'''

            balance_sheet = self.statement

            balance_sheet = self.convert_data(balance_sheet)

            self.download(balance_sheet)
            
            self.merging_data(balance_sheet)
            
        
        elif self.typestatement == 'CF':
            
            sheet = self.get_sheets()
            
            '''Create an Excel Sheet Because there is no in the directory'''

            cash_flow = self.statement

            cash_flow = self.convert_data(cash_flow)

            self.download(cash_flow)
            
            self.merging_data(cash_flow)

        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        